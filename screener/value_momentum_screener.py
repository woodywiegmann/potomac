"""
Value + Momentum Composite Stock Screener (Long/Short + TLH)
=============================================================
QVAL/QMOM-style composite screener for S&P 500 + S&P 400 (~900 stocks).
90% long (top 50) / 10% short (bottom 20) with TLH swap pairs.

Outputs:
  1. Rich Excel workbook (9 sheets) with long book, short book, TLH swaps
  2. HTML dashboard with both books + swap pairs
  3. TradingView watchlist
  4. Console summary

Usage:
    python value_momentum_screener.py                   # default 50/50 V/M
    python value_momentum_screener.py --vw 0.6 --mw 0.4 # 60% value, 40% momentum
    python value_momentum_screener.py --from-config      # read weights from Excel Config sheet
    python value_momentum_screener.py --top 30           # top 30 long basket
    python value_momentum_screener.py --n-short 10       # bottom 10 short basket
"""

import argparse
import io
import json
import os
import sys
import time
import warnings
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta

import numpy as np
import pandas as pd
import requests as _requests
import yfinance as yf

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

warnings.filterwarnings("ignore")

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CACHE_FILE = os.path.join(SCRIPT_DIR, "value_momentum_cache.json")
EXCEL_OUT = os.path.join(SCRIPT_DIR, "value_momentum_dashboard.xlsx")
HTML_OUT = os.path.join(SCRIPT_DIR, "value_momentum_dashboard.html")
TV_OUT = os.path.join(SCRIPT_DIR, "tv_value_momentum_watchlist.txt")

SP500_URL = "https://en.wikipedia.org/wiki/List_of_S%26P_500_companies"
SP400_URL = "https://en.wikipedia.org/wiki/List_of_S%26P_400_companies"

EXCLUDED_SECTORS = {"Financials", "Real Estate"}

LONG_ALLOC = 0.90
SHORT_ALLOC = 0.10

# ── Default weights (overridable via CLI or Config sheet) ─────────────────────

DEFAULT_VALUE_WEIGHTS = {
    "ev_ebit": 0.30,
    "pe": 0.20,
    "pb": 0.15,
    "p_fcf": 0.20,
    "ev_ebitda": 0.15,
}

DEFAULT_MOMENTUM_WEIGHTS = {
    "mom_12_1": 0.35,
    "mom_6": 0.25,
    "mom_3": 0.20,
    "mom_1": 0.10,
    "path_quality": 0.10,
}

DEFAULT_COMPOSITE_WEIGHTS = {"value": 0.50, "momentum": 0.50}

# ── Universe fetch ────────────────────────────────────────────────────────────

_WIKI_HEADERS = {
    "User-Agent": "PotomacScreener/1.0 (woody@potomacfund.com) pandas/read_html"
}


def _fetch_wiki_html(url):
    resp = _requests.get(url, headers=_WIKI_HEADERS, timeout=15)
    resp.raise_for_status()
    return resp.text


def get_sp500_tickers():
    try:
        html = _fetch_wiki_html(SP500_URL)
        tables = pd.read_html(io.StringIO(html))
        df = tables[0]
        tickers = df["Symbol"].str.replace(".", "-", regex=False).tolist()
        sectors = dict(zip(tickers, df["GICS Sector"]))
        names = dict(zip(tickers, df.get("Security", df.get("Company", [""] * len(tickers)))))
        return tickers, sectors, names
    except Exception as e:
        print(f"  WARNING: S&P 500 fetch failed ({e})")
        return [], {}, {}


def get_sp400_tickers():
    try:
        html = _fetch_wiki_html(SP400_URL)
        tables = pd.read_html(io.StringIO(html))
        df = tables[0]
        sym_col = "Symbol" if "Symbol" in df.columns else "Ticker Symbol" if "Ticker Symbol" in df.columns else df.columns[0]
        sect_col = "GICS Sector" if "GICS Sector" in df.columns else "Sector" if "Sector" in df.columns else None
        name_col = "Security" if "Security" in df.columns else "Company" if "Company" in df.columns else None
        tickers = df[sym_col].str.replace(".", "-", regex=False).tolist()
        sectors = dict(zip(tickers, df[sect_col])) if sect_col else {}
        names = dict(zip(tickers, df[name_col])) if name_col else {}
        return tickers, sectors, names
    except Exception as e:
        print(f"  WARNING: S&P 400 fetch failed ({e})")
        return [], {}, {}


def build_universe():
    print("  Fetching S&P 500 tickers...")
    t500, s500, n500 = get_sp500_tickers()
    print(f"    Got {len(t500)} S&P 500 tickers")

    print("  Fetching S&P 400 tickers...")
    t400, s400, n400 = get_sp400_tickers()
    print(f"    Got {len(t400)} S&P 400 tickers")

    all_sectors = {**s500, **s400}
    all_names = {**n500, **n400}
    all_tickers = list(dict.fromkeys(t500 + t400))

    before = len(all_tickers)
    all_tickers = [t for t in all_tickers if all_sectors.get(t, "") not in EXCLUDED_SECTORS]
    print(f"    Excluded {before - len(all_tickers)} Financials/REITs -> {len(all_tickers)} remaining")

    return all_tickers, all_sectors, all_names

# ── Price data ────────────────────────────────────────────────────────────────

def fetch_prices(tickers, lookback_days=380):
    end = datetime.now()
    start = (end - timedelta(days=lookback_days)).strftime("%Y-%m-%d")
    end_str = end.strftime("%Y-%m-%d")
    print(f"  Downloading daily prices ({start} to {end_str})...")
    raw = yf.download(tickers, start=start, end=end_str, auto_adjust=True, progress=False)
    if raw.empty:
        return pd.DataFrame()
    if isinstance(raw.columns, pd.MultiIndex):
        closes = raw["Close"]
    else:
        closes = raw[["Close"]].rename(columns={"Close": tickers[0]})
    closes = closes.ffill()
    print(f"    Got {len(closes)} trading days, {len(closes.columns)} tickers")
    return closes

# ── Momentum metrics ─────────────────────────────────────────────────────────

def compute_momentum(closes):
    results = {}
    daily_rets = closes.pct_change()

    for ticker in closes.columns:
        series = closes[ticker].dropna()
        if len(series) < 252:
            continue
        dr = daily_rets[ticker].dropna()
        price_now = series.iloc[-1]

        mom_12_1 = None
        if len(series) >= 252:
            price_12m = series.iloc[-252]
            price_1m = series.iloc[-21]
            if price_12m > 0 and price_1m > 0:
                mom_12_1 = (price_1m / price_12m) - 1.0

        mom_6 = (price_now / series.iloc[-126] - 1.0) if len(series) >= 126 and series.iloc[-126] > 0 else None
        mom_3 = (price_now / series.iloc[-63] - 1.0) if len(series) >= 63 and series.iloc[-63] > 0 else None
        mom_1 = (price_now / series.iloc[-21] - 1.0) if len(series) >= 21 and series.iloc[-21] > 0 else None

        path_quality = None
        if len(dr) >= 252:
            lookback_rets = dr.iloc[-252:-21]
            if len(lookback_rets) > 0:
                pos_days = (lookback_rets > 0).sum()
                path_quality = pos_days / len(lookback_rets)

        results[ticker] = {
            "mom_12_1": mom_12_1, "mom_6": mom_6, "mom_3": mom_3,
            "mom_1": mom_1, "path_quality": path_quality,
        }

    return pd.DataFrame(results).T

# ── Fundamental data (cached, parallel) ──────────────────────────────────────

def _fetch_single_fundamental(ticker):
    try:
        tk = yf.Ticker(ticker)
        info = tk.info
        if not info or info.get("quoteType") == "NONE":
            return ticker, None

        ev = info.get("enterpriseValue")
        ebit = info.get("ebit")
        ebitda = info.get("ebitda")
        mcap = info.get("marketCap")
        fcf = info.get("freeCashflow")
        gross_profit = info.get("grossProfits")
        total_assets = info.get("totalAssets")

        ev_ebit = ev / ebit if ev and ebit and ebit > 0 else None
        ev_ebitda = ev / ebitda if ev and ebitda and ebitda > 0 else None
        p_fcf = mcap / fcf if mcap and fcf and fcf > 0 else None
        gp_assets = gross_profit / total_assets if gross_profit and total_assets and total_assets > 0 else None

        return ticker, {
            "name": info.get("shortName", info.get("longName", "")),
            "sector": info.get("sector", ""),
            "industry": info.get("industry", ""),
            "mktcap": mcap,
            "mktcap_B": mcap / 1e9 if mcap else None,
            "ev_ebit": ev_ebit,
            "pe": info.get("trailingPE"),
            "forward_pe": info.get("forwardPE"),
            "pb": info.get("priceToBook"),
            "p_fcf": p_fcf,
            "ev_ebitda": ev_ebitda,
            "roe": info.get("returnOnEquity"),
            "debt_equity": info.get("debtToEquity"),
            "gp_assets": gp_assets,
            "gross_margin": info.get("grossMargins"),
            "trailing_eps": info.get("trailingEps"),
            "revenue_growth": info.get("revenueGrowth"),
            "exchange": info.get("exchange", ""),
        }
    except Exception:
        return ticker, None


def fetch_fundamentals(tickers, max_workers=10):
    today_str = datetime.now().strftime("%Y-%m-%d")

    if os.path.isfile(CACHE_FILE):
        try:
            with open(CACHE_FILE, "r") as f:
                cache = json.load(f)
            if cache.get("date") == today_str and cache.get("data"):
                cached = cache["data"]
                print(f"  Loaded {len(cached)} fundamentals from today's cache")
                return pd.DataFrame(cached).T
        except Exception:
            pass

    print(f"  Fetching fundamentals for {len(tickers)} stocks (parallel, {max_workers} workers)...")
    results = {}
    done = 0
    total = len(tickers)
    start_t = time.time()

    with ThreadPoolExecutor(max_workers=max_workers) as pool:
        futures = {pool.submit(_fetch_single_fundamental, t): t for t in tickers}
        for future in as_completed(futures):
            ticker, data = future.result()
            done += 1
            if data:
                results[ticker] = data
            if done % 50 == 0 or done == total:
                elapsed = time.time() - start_t
                rate = done / elapsed if elapsed > 0 else 0
                eta = (total - done) / rate if rate > 0 else 0
                print(f"    {done}/{total} ({len(results)} valid) "
                      f"[{elapsed:.0f}s elapsed, ~{eta:.0f}s remaining]")

    try:
        with open(CACHE_FILE, "w") as f:
            json.dump({"date": today_str, "data": results}, f, default=str)
        print(f"  Cached {len(results)} fundamentals to {CACHE_FILE}")
    except Exception:
        pass

    return pd.DataFrame(results).T

# ── Percentile ranking ────────────────────────────────────────────────────────

def percentile_rank(series, ascending=True):
    valid = series.dropna()
    if len(valid) == 0:
        return pd.Series(np.nan, index=series.index)
    if ascending:
        ranks = valid.rank(ascending=True, pct=True) * 100
        ranks = 100 - ranks
    else:
        ranks = valid.rank(ascending=True, pct=True) * 100
    return ranks.reindex(series.index)

# ── Quality filters ───────────────────────────────────────────────────────────

def apply_quality_filters(fund_df):
    filters = pd.DataFrame(index=fund_df.index)

    roe = pd.to_numeric(fund_df.get("roe"), errors="coerce")
    filters["roe_pass"] = roe.isna() | (roe > 0.05)
    filters["roe_val"] = roe

    gpa = pd.to_numeric(fund_df.get("gp_assets"), errors="coerce")
    filters["gp_assets_pass"] = gpa.isna() | (gpa > 0.20)
    filters["gp_assets_val"] = gpa

    de = pd.to_numeric(fund_df.get("debt_equity"), errors="coerce")
    filters["de_pass"] = de.isna() | (de < 300)
    filters["de_val"] = de

    eps = pd.to_numeric(fund_df.get("trailing_eps"), errors="coerce")
    filters["eps_pass"] = eps.isna() | (eps > 0)
    filters["eps_val"] = eps

    rg = pd.to_numeric(fund_df.get("revenue_growth"), errors="coerce")
    filters["asset_growth_pass"] = rg.isna() | (rg < 0.30)
    filters["asset_growth_val"] = rg

    filters["all_pass"] = (
        filters["roe_pass"] & filters["gp_assets_pass"] &
        filters["de_pass"] & filters["eps_pass"] & filters["asset_growth_pass"]
    )
    return filters

# ── Composite scoring ─────────────────────────────────────────────────────────

def compute_composite(fund_df, mom_df, val_weights, mom_weights, comp_weights):
    merged = fund_df.join(mom_df, how="inner")

    val_pctiles = pd.DataFrame(index=merged.index)
    for metric in ["ev_ebit", "pe", "pb", "p_fcf", "ev_ebitda"]:
        raw = pd.to_numeric(merged.get(metric), errors="coerce")
        val_pctiles[f"{metric}_pctile"] = percentile_rank(raw, ascending=True)

    val_composite = pd.Series(0.0, index=merged.index)
    for metric, weight in val_weights.items():
        col = f"{metric}_pctile"
        if col in val_pctiles.columns:
            val_composite += val_pctiles[col].fillna(50) * weight
    val_pctiles["value_composite"] = val_composite

    mom_pctiles = pd.DataFrame(index=merged.index)
    for metric in ["mom_12_1", "mom_6", "mom_3", "mom_1", "path_quality"]:
        raw = pd.to_numeric(merged.get(metric), errors="coerce")
        mom_pctiles[f"{metric}_pctile"] = percentile_rank(raw, ascending=False)

    mom_composite = pd.Series(0.0, index=merged.index)
    for metric, weight in mom_weights.items():
        col = f"{metric}_pctile"
        if col in mom_pctiles.columns:
            mom_composite += mom_pctiles[col].fillna(50) * weight
    mom_pctiles["momentum_composite"] = mom_composite

    final_composite = (
        comp_weights["value"] * val_composite +
        comp_weights["momentum"] * mom_composite
    )

    result = merged.copy()
    for col in val_pctiles.columns:
        result[col] = val_pctiles[col]
    for col in mom_pctiles.columns:
        result[col] = mom_pctiles[col]
    result["composite_score"] = final_composite
    return result

# ── TLH swap pair generation ─────────────────────────────────────────────────

def generate_tlh_swaps(all_scored, long_tickers, top_n):
    """For each long position, find 2-3 same-sector substitutes ranked 51-150."""
    ranked = all_scored.sort_values("composite_score", ascending=False)
    bench = ranked.iloc[top_n:150]

    swaps = {}
    for ticker in long_tickers:
        if ticker not in all_scored.index:
            continue
        sector = all_scored.loc[ticker].get("sector", "")
        if not sector:
            continue
        same_sector = bench[bench["sector"] == sector]
        candidates = same_sector.index.tolist()[:3]
        if candidates:
            swaps[ticker] = {
                "sector": sector,
                "swap_candidates": candidates,
                "swap_scores": [f"{all_scored.loc[c, 'composite_score']:.1f}" for c in candidates],
            }
    return swaps

# ── Excel workbook builder ────────────────────────────────────────────────────

def _style_header(ws, ncols):
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_align = Alignment(horizontal="center", wrap_text=True)
    thin_border = Border(bottom=Side(style="thin", color="AAAAAA"))
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def _style_header_red(ws, ncols):
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    header_align = Alignment(horizontal="center", wrap_text=True)
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align


def _auto_width(ws, max_width=25):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, max_width)


def _add_color_scale(ws, col_idx, min_row, max_row):
    col_letter = get_column_letter(col_idx)
    cell_range = f"{col_letter}{min_row}:{col_letter}{max_row}"
    rule = ColorScaleRule(
        start_type="min", start_color="F8696B",
        mid_type="percentile", mid_value=50, mid_color="FFEB84",
        end_type="max", end_color="63BE7B",
    )
    ws.conditional_formatting.add(cell_range, rule)


def _df_to_sheet(wb, sheet_name, df, pctile_cols=None, red_header=False):
    ws = wb.create_sheet(title=sheet_name)
    headers = list(df.columns)
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)

    num_fmt_1f = '0.0'
    num_fmt_2f = '0.00'
    num_fmt_0f = '#,##0'

    for r, (idx, row) in enumerate(df.iterrows(), 2):
        for c, col in enumerate(headers, 1):
            val = row[col]
            cell = ws.cell(row=r, column=c)
            if isinstance(val, (np.floating, float)):
                if np.isnan(val) or np.isinf(val):
                    cell.value = None
                else:
                    cell.value = float(val)
                    if "pctile" in col or col in ("value_composite", "momentum_composite", "composite_score"):
                        cell.number_format = num_fmt_1f
                    elif "mom_" in col or col == "path_quality" or "roe" in col or "margin" in col or "growth" in col:
                        cell.number_format = '0.00%'
                    elif col == "mktcap_B":
                        cell.number_format = num_fmt_1f
                    elif col in ("ev_ebit", "pe", "pb", "p_fcf", "ev_ebitda", "forward_pe", "debt_equity", "weight"):
                        cell.number_format = num_fmt_1f
                    else:
                        cell.number_format = num_fmt_2f
            elif isinstance(val, (np.integer, int)):
                cell.value = int(val)
                cell.number_format = num_fmt_0f
            elif isinstance(val, (bool, np.bool_)):
                cell.value = "PASS" if val else "FAIL"
                cell.font = Font(color="2E7D32") if val else Font(color="CC0000", bold=True)
            else:
                cell.value = str(val) if val is not None else ""

    if red_header:
        _style_header_red(ws, len(headers))
    else:
        _style_header(ws, len(headers))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _auto_width(ws)

    if pctile_cols:
        max_row = len(df) + 1
        for col_name in pctile_cols:
            if col_name in headers:
                col_idx = headers.index(col_name) + 1
                _add_color_scale(ws, col_idx, 2, max_row)

    return ws


def build_excel(all_scored, quality_filters, long_df, short_df, tlh_swaps,
                val_weights, mom_weights, comp_weights, top_n, n_short):
    if not HAS_OPENPYXL:
        print("  WARNING: openpyxl not installed, skipping Excel output")
        return

    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: All Stocks ───────────────────────────────────────────────
    all_cols = ["name", "sector", "industry", "mktcap_B", "exchange",
                "ev_ebit", "pe", "forward_pe", "pb", "p_fcf", "ev_ebitda",
                "roe", "debt_equity", "gp_assets", "gross_margin", "trailing_eps", "revenue_growth",
                "mom_12_1", "mom_6", "mom_3", "mom_1", "path_quality",
                "ev_ebit_pctile", "pe_pctile", "pb_pctile", "p_fcf_pctile", "ev_ebitda_pctile",
                "value_composite",
                "mom_12_1_pctile", "mom_6_pctile", "mom_3_pctile", "mom_1_pctile", "path_quality_pctile",
                "momentum_composite", "composite_score"]
    available = [c for c in all_cols if c in all_scored.columns]
    sheet_df = all_scored[available].copy()
    sheet_df.insert(0, "ticker", sheet_df.index)
    sheet_df = sheet_df.sort_values("composite_score", ascending=False)
    pctile_cols = [c for c in available if "pctile" in c or c in ("value_composite", "momentum_composite", "composite_score")]
    _df_to_sheet(wb, "All Stocks", sheet_df, pctile_cols=pctile_cols)

    # ── Sheet 2: Value Rankings ───────────────────────────────────────────
    val_cols = ["name", "sector", "mktcap_B",
                "ev_ebit", "ev_ebit_pctile", "pe", "pe_pctile", "pb", "pb_pctile",
                "p_fcf", "p_fcf_pctile", "ev_ebitda", "ev_ebitda_pctile", "value_composite"]
    avail_val = [c for c in val_cols if c in all_scored.columns]
    val_df = all_scored[avail_val].copy()
    val_df.insert(0, "ticker", val_df.index)
    val_df = val_df.sort_values("value_composite", ascending=False)
    val_df.insert(1, "value_rank", range(1, len(val_df) + 1))
    _df_to_sheet(wb, "Value Rankings", val_df,
                 pctile_cols=[c for c in avail_val if "pctile" in c or c == "value_composite"])

    # ── Sheet 3: Momentum Rankings ────────────────────────────────────────
    mom_cols = ["name", "sector", "mktcap_B",
                "mom_12_1", "mom_12_1_pctile", "mom_6", "mom_6_pctile",
                "mom_3", "mom_3_pctile", "mom_1", "mom_1_pctile",
                "path_quality", "path_quality_pctile", "momentum_composite"]
    avail_mom = [c for c in mom_cols if c in all_scored.columns]
    mom_df = all_scored[avail_mom].copy()
    mom_df.insert(0, "ticker", mom_df.index)
    mom_df = mom_df.sort_values("momentum_composite", ascending=False)
    mom_df.insert(1, "momentum_rank", range(1, len(mom_df) + 1))
    _df_to_sheet(wb, "Momentum Rankings", mom_df,
                 pctile_cols=[c for c in avail_mom if "pctile" in c or c == "momentum_composite"])

    # ── Sheet 4: Quality Screen ───────────────────────────────────────────
    qf = quality_filters.copy()
    qf.insert(0, "ticker", qf.index)
    if "name" in all_scored.columns:
        qf.insert(1, "name", all_scored["name"].reindex(qf.index))
    if "sector" in all_scored.columns:
        qf.insert(2, "sector", all_scored["sector"].reindex(qf.index))
    _df_to_sheet(wb, "Quality Screen", qf)

    # ── Sheet 5: Composite Rankings ───────────────────────────────────────
    comp_cols = ["name", "sector", "mktcap_B", "value_composite", "momentum_composite", "composite_score"]
    avail_comp = [c for c in comp_cols if c in all_scored.columns]
    comp_df = all_scored[avail_comp].copy()
    comp_df.insert(0, "ticker", comp_df.index)
    comp_df = comp_df.sort_values("composite_score", ascending=False)
    comp_df.insert(1, "composite_rank", range(1, len(comp_df) + 1))
    _df_to_sheet(wb, "Composite Rankings", comp_df,
                 pctile_cols=["value_composite", "momentum_composite", "composite_score"])

    # ── Sheet 6: Long Book (Top N) ────────────────────────────────────────
    basket_cols = ["name", "sector", "mktcap_B", "ev_ebit", "pe", "pb", "p_fcf",
                   "mom_12_1", "mom_6", "mom_3", "path_quality",
                   "value_composite", "momentum_composite", "composite_score"]
    avail_basket = [c for c in basket_cols if c in long_df.columns]
    basket_df = long_df[avail_basket].copy()
    basket_df.insert(0, "ticker", basket_df.index)
    basket_df.insert(1, "rank", range(1, len(basket_df) + 1))
    basket_df["weight"] = LONG_ALLOC / len(basket_df)

    ws = _df_to_sheet(wb, f"Long Book ({top_n})", basket_df,
                      pctile_cols=["value_composite", "momentum_composite", "composite_score"])

    sr = len(basket_df) + 3
    ws.cell(row=sr, column=1, value="LONG BOOK SUMMARY").font = Font(bold=True, size=12, color="1F4E79")
    ws.cell(row=sr+1, column=1, value="Stocks"); ws.cell(row=sr+1, column=2, value=len(basket_df))
    ws.cell(row=sr+2, column=1, value="Allocation"); ws.cell(row=sr+2, column=2, value=f"{LONG_ALLOC:.0%}")
    ws.cell(row=sr+3, column=1, value="Per-stock weight"); ws.cell(row=sr+3, column=2, value=f"{LONG_ALLOC/len(basket_df):.2%}")
    ws.cell(row=sr+4, column=1, value="Avg Market Cap ($B)"); ws.cell(row=sr+4, column=2, value=round(pd.to_numeric(basket_df["mktcap_B"], errors="coerce").mean(), 1))
    ws.cell(row=sr+5, column=1, value="Avg Composite"); ws.cell(row=sr+5, column=2, value=round(basket_df["composite_score"].mean(), 1))

    if "sector" in basket_df.columns:
        ws.cell(row=sr+7, column=1, value="SECTOR BREAKDOWN").font = Font(bold=True, size=11, color="1F4E79")
        for i, (sect, cnt) in enumerate(basket_df["sector"].value_counts().items()):
            ws.cell(row=sr+8+i, column=1, value=sect)
            ws.cell(row=sr+8+i, column=2, value=cnt)
            ws.cell(row=sr+8+i, column=3, value=f"{cnt/len(basket_df):.0%}")

    # ── Sheet 7: Short Book (Bottom N) ────────────────────────────────────
    short_cols = ["name", "sector", "mktcap_B", "ev_ebit", "pe", "pb",
                  "mom_12_1", "mom_6", "mom_3",
                  "value_composite", "momentum_composite", "composite_score"]
    avail_short = [c for c in short_cols if c in short_df.columns]
    s_df = short_df[avail_short].copy()
    s_df.insert(0, "ticker", s_df.index)
    s_df.insert(1, "rank", range(1, len(s_df) + 1))
    s_df["weight"] = SHORT_ALLOC / len(s_df)

    ws_s = _df_to_sheet(wb, f"Short Book ({n_short})", s_df, red_header=True,
                        pctile_cols=["value_composite", "momentum_composite", "composite_score"])

    sr = len(s_df) + 3
    ws_s.cell(row=sr, column=1, value="SHORT BOOK SUMMARY").font = Font(bold=True, size=12, color="8B0000")
    ws_s.cell(row=sr+1, column=1, value="Stocks"); ws_s.cell(row=sr+1, column=2, value=len(s_df))
    ws_s.cell(row=sr+2, column=1, value="Allocation"); ws_s.cell(row=sr+2, column=2, value=f"{SHORT_ALLOC:.0%}")
    ws_s.cell(row=sr+3, column=1, value="Per-stock weight"); ws_s.cell(row=sr+3, column=2, value=f"{SHORT_ALLOC/len(s_df):.2%}")
    ws_s.cell(row=sr+4, column=1, value="Avg Composite (lower=worse)"); ws_s.cell(row=sr+4, column=2, value=round(s_df["composite_score"].mean(), 1))

    # ── Sheet 8: TLH Swaps ────────────────────────────────────────────────
    swap_rows = []
    for ticker, info in tlh_swaps.items():
        cands = info["swap_candidates"]
        scores = info["swap_scores"]
        for i, (c, s) in enumerate(zip(cands, scores)):
            swap_rows.append({
                "primary_ticker": ticker,
                "sector": info["sector"],
                "swap_rank": i + 1,
                "swap_ticker": c,
                "swap_composite": float(s),
            })
    if swap_rows:
        swap_df = pd.DataFrame(swap_rows)
        _df_to_sheet(wb, "TLH Swaps", swap_df)
    else:
        ws_tlh = wb.create_sheet(title="TLH Swaps")
        ws_tlh.cell(row=1, column=1, value="No swap pairs generated")

    # ── Sheet 9: Config ───────────────────────────────────────────────────
    ws_cfg = wb.create_sheet(title="Config")
    ws_cfg.cell(row=1, column=1, value="VALUE + MOMENTUM SCREENER CONFIGURATION").font = Font(bold=True, size=14, color="1F4E79")
    ws_cfg.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(color="666666")
    ws_cfg.cell(row=3, column=1, value="Edit weights below, then re-run with: python value_momentum_screener.py --from-config").font = Font(italic=True, color="666666")

    ws_cfg.cell(row=5, column=1, value="COMPOSITE BLEND").font = Font(bold=True, size=12, color="2C6E49")
    ws_cfg.cell(row=6, column=1, value="Value Weight"); ws_cfg.cell(row=6, column=2, value=comp_weights["value"])
    ws_cfg.cell(row=7, column=1, value="Momentum Weight"); ws_cfg.cell(row=7, column=2, value=comp_weights["momentum"])

    ws_cfg.cell(row=9, column=1, value="VALUE SUB-FACTOR WEIGHTS").font = Font(bold=True, size=12, color="2C6E49")
    row = 10
    for metric, weight in val_weights.items():
        ws_cfg.cell(row=row, column=1, value=metric); ws_cfg.cell(row=row, column=2, value=weight)
        row += 1

    row += 1
    ws_cfg.cell(row=row, column=1, value="MOMENTUM SUB-FACTOR WEIGHTS").font = Font(bold=True, size=12, color="2C6E49")
    row += 1
    for metric, weight in mom_weights.items():
        ws_cfg.cell(row=row, column=1, value=metric); ws_cfg.cell(row=row, column=2, value=weight)
        row += 1

    row += 1
    ws_cfg.cell(row=row, column=1, value="BASKET SIZE").font = Font(bold=True, size=12, color="2C6E49")
    row += 1
    ws_cfg.cell(row=row, column=1, value="top_n"); ws_cfg.cell(row=row, column=2, value=top_n)
    row += 1
    ws_cfg.cell(row=row, column=1, value="n_short"); ws_cfg.cell(row=row, column=2, value=n_short)

    _auto_width(ws_cfg)
    wb.save(EXCEL_OUT)
    print(f"\n  Excel workbook saved to: {EXCEL_OUT}")

# ── Config reader ─────────────────────────────────────────────────────────────

def read_config_from_excel():
    if not os.path.isfile(EXCEL_OUT):
        print(f"  ERROR: {EXCEL_OUT} not found. Run without --from-config first.")
        sys.exit(1)

    wb = load_workbook(EXCEL_OUT, data_only=True)
    ws = wb["Config"]

    comp_weights = {
        "value": float(ws.cell(row=6, column=2).value or 0.5),
        "momentum": float(ws.cell(row=7, column=2).value or 0.5),
    }

    val_weights = {}
    row = 10
    for metric in DEFAULT_VALUE_WEIGHTS:
        val_weights[metric] = float(ws.cell(row=row, column=2).value or DEFAULT_VALUE_WEIGHTS[metric])
        row += 1

    mom_weights = {}
    row += 2
    for metric in DEFAULT_MOMENTUM_WEIGHTS:
        mom_weights[metric] = float(ws.cell(row=row, column=2).value or DEFAULT_MOMENTUM_WEIGHTS[metric])
        row += 1

    row += 2
    top_n = int(ws.cell(row=row, column=2).value or 50)
    row += 1
    n_short = int(ws.cell(row=row, column=2).value or 20)

    wb.close()
    return val_weights, mom_weights, comp_weights, top_n, n_short

# ── TradingView export ────────────────────────────────────────────────────────

EXCHANGE_MAP = {
    "NMS": "NASDAQ", "NGM": "NASDAQ", "NCM": "NASDAQ", "NAS": "NASDAQ",
    "NYQ": "NYSE", "NYS": "NYSE", "PCX": "NYSE", "ASE": "NYSE",
    "BTS": "NYSE", "CBO": "NYSE",
}


def export_tradingview(all_data, top_n):
    top = all_data.sort_values("composite_score", ascending=False).head(top_n)
    lines = []
    plain_tickers = []
    for ticker in top.index:
        exch_raw = str(top.loc[ticker].get("exchange", "")).upper()
        exch = EXCHANGE_MAP.get(exch_raw, "NYSE")
        lines.append(f"{exch}:{ticker}")
        plain_tickers.append(ticker)

    with open(TV_OUT, "w") as f:
        f.write("\n".join(lines))

    print(f"\n  TradingView watchlist saved to: {TV_OUT}")
    print(f"  Paste into TradingView: {', '.join(plain_tickers)}")

# ── HTML dashboard ────────────────────────────────────────────────────────────

def _html_stock_row(i, ticker, row, side="long"):
    name = str(row.get("name", ""))[:25]
    sector = str(row.get("sector", ""))[:20]
    mcap = f"${row.get('mktcap_B', 0):.0f}B" if pd.notna(row.get("mktcap_B")) else "n/a"
    vscore = f"{row.get('value_composite', 0):.1f}" if pd.notna(row.get("value_composite")) else "n/a"
    mscore = f"{row.get('momentum_composite', 0):.1f}" if pd.notna(row.get("momentum_composite")) else "n/a"
    comp = f"{row.get('composite_score', 0):.1f}" if pd.notna(row.get("composite_score")) else "n/a"
    ev_ebit = f"{row.get('ev_ebit', 0):.1f}" if pd.notna(row.get("ev_ebit")) else "n/a"
    pe = f"{row.get('pe', 0):.1f}" if pd.notna(row.get("pe")) else "n/a"
    mom12 = f"{row.get('mom_12_1', 0):.1%}" if pd.notna(row.get("mom_12_1")) else "n/a"
    mom6 = f"{row.get('mom_6', 0):.1%}" if pd.notna(row.get("mom_6")) else "n/a"

    def sc(val_str):
        try:
            v = float(val_str)
            if v >= 70: return "score-high"
            if v >= 40: return "score-mid"
            return "score-low"
        except (ValueError, TypeError):
            return ""

    ticker_cls = "ticker" if side == "long" else "ticker-short"
    return (f"    <tr><td>{i+1}</td><td class=\"{ticker_cls}\"><b>{ticker}</b></td>"
            f"<td>{name}</td><td>{sector}</td><td>{mcap}</td>"
            f"<td>{ev_ebit}</td><td>{pe}</td><td>{mom12}</td><td>{mom6}</td>"
            f"<td class=\"{sc(vscore)}\">{vscore}</td>"
            f"<td class=\"{sc(mscore)}\">{mscore}</td>"
            f"<td class=\"{sc(comp)}\"><b>{comp}</b></td></tr>")


def build_html_dashboard(all_scored, quality_filters, long_df, short_df, tlh_swaps, comp_weights, top_n, n_short):
    as_of = datetime.now().strftime("%Y-%m-%d %H:%M")
    total_screened = len(all_scored)
    quality_pass = len(quality_filters[quality_filters["all_pass"]]) if "all_pass" in quality_filters.columns else total_screened

    long_rows = "\n".join([_html_stock_row(i, t, long_df.loc[t]) for i, t in enumerate(long_df.index)])
    short_rows = "\n".join([_html_stock_row(i, t, short_df.loc[t], "short") for i, t in enumerate(short_df.index)])

    sector_rows = []
    if "sector" in long_df.columns:
        for sect, cnt in long_df["sector"].value_counts().items():
            pct = cnt / len(long_df) * 100
            sector_rows.append(f'    <tr><td>{sect}</td><td>{cnt}</td><td>{pct:.0f}%</td>'
                               f'<td><div class="bar" style="width:{pct*3}px"></div></td></tr>')
    sector_table = "\n".join(sector_rows)

    avg_mcap = pd.to_numeric(long_df.get("mktcap_B"), errors="coerce").mean()
    avg_comp = long_df["composite_score"].mean() if "composite_score" in long_df.columns else 0
    short_avg = short_df["composite_score"].mean() if "composite_score" in short_df.columns else 0

    tv_tickers = ", ".join(long_df.index.tolist())

    swap_rows_html = []
    for ticker, info in tlh_swaps.items():
        cands = ", ".join([f"{c} ({s})" for c, s in zip(info["swap_candidates"], info["swap_scores"])])
        swap_rows_html.append(f'    <tr><td class="ticker"><b>{ticker}</b></td><td>{info["sector"]}</td><td>{cands}</td></tr>')
    swap_table = "\n".join(swap_rows_html) if swap_rows_html else '    <tr><td colspan="3">No swap pairs available</td></tr>'

    tbl_header = "<tr><th>#</th><th>Ticker</th><th>Name</th><th>Sector</th><th>Mkt Cap</th><th>EV/EBIT</th><th>P/E</th><th>Mom 12-1</th><th>Mom 6m</th><th>Value</th><th>Mom</th><th>Composite</th></tr>"

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <title>Value + Momentum Composite Dashboard</title>
  <style>
    * {{ box-sizing: border-box; margin: 0; padding: 0; }}
    body {{ font-family: 'Segoe UI', system-ui, -apple-system, sans-serif; background: #0f1117; color: #e1e4e8; padding: 20px; }}
    .header {{ background: linear-gradient(135deg, #1f4e79 0%, #2c6e49 100%); padding: 24px 32px; border-radius: 12px; margin-bottom: 20px; }}
    .header h1 {{ color: #fff; font-size: 1.8em; margin-bottom: 4px; }}
    .header .sub {{ color: #b0d4c8; font-size: 0.95em; }}
    .cards {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(160px, 1fr)); gap: 12px; margin-bottom: 20px; }}
    .card {{ background: #161b22; border: 1px solid #30363d; border-radius: 10px; padding: 16px; text-align: center; }}
    .card .label {{ color: #8b949e; font-size: 0.8em; text-transform: uppercase; letter-spacing: 0.5px; }}
    .card .value {{ font-size: 1.5em; font-weight: 700; color: #58a6ff; margin-top: 4px; }}
    .card .value.green {{ color: #3fb950; }}
    .card .value.red {{ color: #f85149; }}
    .section {{ background: #161b22; border: 1px solid #30363d; border-radius: 10px; padding: 20px; margin-bottom: 20px; }}
    .section h2 {{ color: #58a6ff; font-size: 1.15em; margin-bottom: 12px; border-bottom: 1px solid #30363d; padding-bottom: 8px; }}
    .section h2.short {{ color: #f85149; }}
    .section h2.tlh {{ color: #d29922; }}
    table {{ width: 100%; border-collapse: collapse; font-size: 0.85em; }}
    th {{ background: #1f2937; color: #8b949e; padding: 8px 10px; text-align: left; font-weight: 600;
         text-transform: uppercase; font-size: 0.75em; letter-spacing: 0.5px; position: sticky; top: 0; }}
    td {{ padding: 6px 10px; border-bottom: 1px solid #21262d; }}
    tr:hover {{ background: #1c2333; }}
    .ticker {{ color: #58a6ff; }}
    .ticker-short {{ color: #f85149; }}
    .score-high {{ color: #3fb950; font-weight: 600; }}
    .score-mid {{ color: #d29922; }}
    .score-low {{ color: #f85149; }}
    .bar {{ background: linear-gradient(90deg, #1f4e79, #58a6ff); height: 14px; border-radius: 3px; }}
    .tv-box {{ background: #0d1117; border: 1px solid #30363d; border-radius: 8px; padding: 12px 16px;
               font-family: 'Consolas', 'Courier New', monospace; font-size: 0.82em; color: #8b949e;
               word-break: break-all; line-height: 1.6; margin-top: 8px; }}
    .grid-2 {{ display: grid; grid-template-columns: 1fr 1fr; gap: 20px; }}
    @media (max-width: 900px) {{ .grid-2 {{ grid-template-columns: 1fr; }} }}
    .meta {{ color: #484f58; font-size: 0.78em; text-align: center; margin-top: 16px; }}
    .scroll-table {{ max-height: 60vh; overflow-y: auto; }}
  </style>
</head>
<body>

<div class="header">
  <h1>Value + Momentum Composite (90/10 Long/Short)</h1>
  <div class="sub">QVAL/QMOM-style &bull; S&amp;P 500 + 400 &bull; Quarterly rebalance &bull; As of {as_of}</div>
</div>

<div class="cards">
  <div class="card"><div class="label">Blend</div><div class="value">{comp_weights['value']:.0%}V / {comp_weights['momentum']:.0%}M</div></div>
  <div class="card"><div class="label">Universe</div><div class="value">{total_screened}</div></div>
  <div class="card"><div class="label">Quality Pass</div><div class="value green">{quality_pass}</div></div>
  <div class="card"><div class="label">Long Book</div><div class="value green">{top_n} @ {LONG_ALLOC:.0%}</div></div>
  <div class="card"><div class="label">Short Book</div><div class="value red">{n_short} @ {SHORT_ALLOC:.0%}</div></div>
  <div class="card"><div class="label">Net Exposure</div><div class="value">{LONG_ALLOC - SHORT_ALLOC:.0%}</div></div>
  <div class="card"><div class="label">Avg Mkt Cap</div><div class="value">${avg_mcap:.0f}B</div></div>
  <div class="card"><div class="label">TLH Pairs</div><div class="value">{len(tlh_swaps)}</div></div>
</div>

<div class="section">
  <h2>Long Book: Top {top_n} (90% allocation, ~{LONG_ALLOC/top_n:.1%} each)</h2>
  <div class="scroll-table">
  <table><thead>{tbl_header}</thead><tbody>
{long_rows}
  </tbody></table></div>
</div>

<div class="section">
  <h2 class="short">Short Book: Bottom {n_short} (10% allocation, ~{SHORT_ALLOC/n_short:.1%} each)</h2>
  <div class="scroll-table">
  <table><thead>{tbl_header}</thead><tbody>
{short_rows}
  </tbody></table></div>
</div>

<div class="grid-2">
  <div class="section">
    <h2>Long Book Sectors</h2>
    <table><thead><tr><th>Sector</th><th>Count</th><th>Weight</th><th></th></tr></thead><tbody>
{sector_table}
    </tbody></table>
  </div>
  <div class="section">
    <h2>Portfolio Statistics</h2>
    <table><tbody>
      <tr><td>Long Avg Composite</td><td class="score-high">{avg_comp:.1f}</td></tr>
      <tr><td>Short Avg Composite</td><td class="score-low">{short_avg:.1f}</td></tr>
      <tr><td>Long/Short Spread</td><td class="score-high">{avg_comp - short_avg:.1f} pts</td></tr>
      <tr><td>Long Avg Mkt Cap</td><td>${avg_mcap:.1f}B</td></tr>
      <tr><td>Gross Exposure</td><td>{LONG_ALLOC + SHORT_ALLOC:.0%}</td></tr>
      <tr><td>Net Exposure</td><td>{LONG_ALLOC - SHORT_ALLOC:.0%}</td></tr>
      <tr><td>Rebalance</td><td>Quarterly (Mar/Jun/Sep/Dec)</td></tr>
    </tbody></table>
  </div>
</div>

<div class="section">
  <h2 class="tlh">TLH Swap Pairs (same-sector substitutes for long book)</h2>
  <div class="scroll-table">
  <table><thead><tr><th>Primary</th><th>Sector</th><th>Swap Candidates (composite score)</th></tr></thead><tbody>
{swap_table}
  </tbody></table></div>
</div>

<div class="section">
  <h2>TradingView Paste (long book)</h2>
  <div class="tv-box">{tv_tickers}</div>
</div>

<div class="meta">
  Generated by value_momentum_screener.py &bull; Potomac Fund Management &bull; {as_of}<br>
  90/10 Long/Short &bull; Quarterly rebalance &bull; <code>python value_momentum_screener.py</code>
</div>

</body>
</html>"""

    with open(HTML_OUT, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"\n  HTML dashboard saved to: {HTML_OUT}")

# ── Console output ────────────────────────────────────────────────────────────

def _print_book(df, title, weight_per_stock):
    print(f"\n{'=' * 95}")
    print(f"  {title}")
    print(f"{'=' * 95}")
    print(f"  {'Rank':<5} {'Ticker':<7} {'Name':<22} {'Sector':<20} {'MktCap':>7} {'V-Score':>8} {'M-Score':>8} {'Comp':>8} {'Wt':>6}")
    print(f"  {'-' * 93}")

    for i, (ticker, row) in enumerate(df.iterrows()):
        name = str(row.get("name", ""))[:20]
        sector = str(row.get("sector", ""))[:18]
        mktcap = f"{row.get('mktcap_B', 0):.0f}B" if pd.notna(row.get("mktcap_B")) else "n/a"
        vscore = f"{row.get('value_composite', 0):.1f}" if pd.notna(row.get("value_composite")) else "n/a"
        mscore = f"{row.get('momentum_composite', 0):.1f}" if pd.notna(row.get("momentum_composite")) else "n/a"
        comp = f"{row.get('composite_score', 0):.1f}" if pd.notna(row.get("composite_score")) else "n/a"
        wt = f"{weight_per_stock:.1%}"
        print(f"  {i+1:<5} {ticker:<7} {name:<22} {sector:<20} {mktcap:>7} {vscore:>8} {mscore:>8} {comp:>8} {wt:>6}")


def print_results(long_df, short_df, tlh_swaps, comp_weights, top_n, n_short):
    _print_book(long_df,
                f"LONG BOOK: TOP {top_n} ({LONG_ALLOC:.0%} allocation, {comp_weights['value']:.0%}V/{comp_weights['momentum']:.0%}M)",
                LONG_ALLOC / top_n)

    print(f"\n  LONG BOOK STATS:")
    print(f"    Avg Composite:  {long_df['composite_score'].mean():.1f}")
    print(f"    Avg Market Cap: ${pd.to_numeric(long_df['mktcap_B'], errors='coerce').mean():.1f}B")
    if "sector" in long_df.columns:
        print(f"  SECTOR BREAKDOWN:")
        for sect, cnt in long_df["sector"].value_counts().items():
            print(f"    {sect:<25} {cnt:>3} ({cnt/len(long_df):.0%})")

    _print_book(short_df,
                f"SHORT BOOK: BOTTOM {n_short} ({SHORT_ALLOC:.0%} allocation)",
                SHORT_ALLOC / n_short)

    print(f"\n  SHORT BOOK STATS:")
    print(f"    Avg Composite:  {short_df['composite_score'].mean():.1f}")
    print(f"    L/S Spread:     {long_df['composite_score'].mean() - short_df['composite_score'].mean():.1f} pts")

    print(f"\n  TLH SWAP PAIRS ({len(tlh_swaps)} of {top_n} long positions have swaps):")
    for ticker, info in list(tlh_swaps.items())[:10]:
        cands = ", ".join(info["swap_candidates"])
        print(f"    {ticker:<7} -> {cands}")
    if len(tlh_swaps) > 10:
        print(f"    ... and {len(tlh_swaps) - 10} more (see Excel/HTML)")

    print(f"\n  PORTFOLIO SUMMARY:")
    print(f"    Gross Exposure: {LONG_ALLOC + SHORT_ALLOC:.0%}")
    print(f"    Net Exposure:   {LONG_ALLOC - SHORT_ALLOC:.0%}")
    print(f"    Rebalance:      Quarterly (Mar/Jun/Sep/Dec)")
    print(f"\n{'=' * 95}")

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Value + Momentum Composite Screener (Long/Short)")
    parser.add_argument("--vw", type=float, default=None, help="Value weight (0-1)")
    parser.add_argument("--mw", type=float, default=None, help="Momentum weight (0-1)")
    parser.add_argument("--top", type=int, default=50, help="Long basket size")
    parser.add_argument("--n-short", type=int, default=20, help="Short basket size")
    parser.add_argument("--from-config", action="store_true", help="Read weights from Excel Config sheet")
    parser.add_argument("--workers", type=int, default=10, help="Parallel workers for fundamental fetch")
    parser.add_argument("--no-excel", action="store_true", help="Skip Excel output")
    args = parser.parse_args()

    val_weights = DEFAULT_VALUE_WEIGHTS.copy()
    mom_weights = DEFAULT_MOMENTUM_WEIGHTS.copy()
    comp_weights = DEFAULT_COMPOSITE_WEIGHTS.copy()
    top_n = args.top
    n_short = args.n_short

    if args.from_config:
        print("  Reading weights from Excel Config sheet...")
        val_weights, mom_weights, comp_weights, top_n, n_short = read_config_from_excel()
        print(f"    Blend: {comp_weights['value']:.0%}/{comp_weights['momentum']:.0%}, Long: {top_n}, Short: {n_short}")
    elif args.vw is not None and args.mw is not None:
        comp_weights = {"value": args.vw, "momentum": args.mw}
    elif args.vw is not None:
        comp_weights = {"value": args.vw, "momentum": 1.0 - args.vw}
    elif args.mw is not None:
        comp_weights = {"value": 1.0 - args.mw, "momentum": args.mw}

    print("=" * 70)
    print("  VALUE + MOMENTUM COMPOSITE SCREENER (90/10 Long/Short)")
    print(f"  Blend: {comp_weights['value']:.0%} Value / {comp_weights['momentum']:.0%} Momentum")
    print(f"  Long: Top {top_n} ({LONG_ALLOC:.0%}) | Short: Bottom {n_short} ({SHORT_ALLOC:.0%})")
    print("=" * 70)

    # 1. Build universe
    print("\n[1/6] Building universe...")
    tickers, sector_map, name_map = build_universe()
    if not tickers:
        print("  ERROR: No tickers found")
        return

    # 2. Fetch prices
    print("\n[2/6] Fetching price data...")
    closes = fetch_prices(tickers, lookback_days=380)
    if closes.empty:
        print("  ERROR: No price data")
        return

    # 3. Compute momentum
    print("\n[3/6] Computing momentum metrics...")
    mom_df = compute_momentum(closes)
    print(f"    Computed momentum for {len(mom_df)} stocks")

    # 4. Fetch fundamentals
    print("\n[4/6] Fetching fundamental data...")
    valid_tickers = [t for t in tickers if t in mom_df.index]
    fund_df = fetch_fundamentals(valid_tickers, max_workers=args.workers)
    print(f"    Got fundamentals for {len(fund_df)} stocks")

    for t in fund_df.index:
        if pd.isna(fund_df.loc[t].get("sector")) or fund_df.loc[t].get("sector") == "":
            if t in sector_map:
                fund_df.at[t, "sector"] = sector_map[t]
        if pd.isna(fund_df.loc[t].get("name")) or fund_df.loc[t].get("name") == "":
            if t in name_map:
                fund_df.at[t, "name"] = name_map[t]

    # 5. Score ALL stocks (needed for both long and short books)
    print("\n[5/6] Computing composite scores for full universe...")
    all_scored = compute_composite(fund_df, mom_df, val_weights, mom_weights, comp_weights)
    print(f"    Scored {len(all_scored)} stocks")

    quality = apply_quality_filters(fund_df)

    # 6. Build long and short books
    print("\n[6/6] Building long/short books and TLH swaps...")

    # Long book: top N from quality-passing stocks only
    passing = quality[quality["all_pass"]].index
    long_candidates = all_scored.loc[all_scored.index.isin(passing)]
    long_df = long_candidates.sort_values("composite_score", ascending=False).head(top_n)
    print(f"    Long book: {len(long_df)} stocks (from {len(long_candidates)} quality-passing)")

    # Short book: bottom N from ALL scored stocks (no quality filter -- bad quality is the point)
    short_df = all_scored.sort_values("composite_score", ascending=True).head(n_short)
    print(f"    Short book: {len(short_df)} stocks (worst composite scores)")

    # TLH swap pairs
    tlh_swaps = generate_tlh_swaps(all_scored, long_df.index.tolist(), top_n)
    print(f"    TLH swaps: {len(tlh_swaps)} of {top_n} positions have same-sector substitutes")

    # Output
    print_results(long_df, short_df, tlh_swaps, comp_weights, top_n, n_short)
    export_tradingview(all_scored, top_n)
    build_html_dashboard(all_scored, quality, long_df, short_df, tlh_swaps, comp_weights, top_n, n_short)

    if not args.no_excel:
        build_excel(all_scored, quality, long_df, short_df, tlh_swaps,
                    val_weights, mom_weights, comp_weights, top_n, n_short)

    print("\nDONE. Outputs:")
    print(f"  Dashboard:  {HTML_OUT}")
    print(f"  Excel:      {EXCEL_OUT}")
    print(f"  Watchlist:  {TV_OUT}")


if __name__ == "__main__":
    main()
